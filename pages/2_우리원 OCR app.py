import streamlit as st
import pandas as pd
import fitz
import pytesseract
from PIL import Image, ImageOps, ImageEnhance
import io
import re
from datetime import datetime
from openpyxl.styles import PatternFill

# --- [설정] Tesseract 경로 (서버 환경에 맞게 수정) ---
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

ORDER_CODE_MAP = {
    "LZCMARC201": "마크로젠 20종(남)",
    "LZCMARC202": "마크로젠 20종(여)",
    "LZCMARCFB": "마크로젠 일반질환 21종(여)",
    "LZCMARCAUTO": "마크로젠 면역질환 16종",
    "LZCMARCMB": "마크로젠 일반질환 19종(남)",
}

class PDFConverter:
    def extract_refined_info(self, original_line):
        line_compact = original_line.replace(' ', '')
        
        # --- 1. 차트번호 추출 (격리 보호 기법) ---
        chart_no = ""
        safe_string = original_line
        for bc in re.findall(r'(26\d{8})', line_compact): safe_string = safe_string.replace(bc, '_')
        for kv in re.findall(r'(2026\d{8})', line_compact): safe_string = safe_string.replace(kv, '_')
        safe_compact = safe_string.replace(' ', '')
        
        potential_charts = re.findall(r'([12]\d{7})', safe_compact)
        for pc in potential_charts:
            if pc.startswith('19') or pc.startswith('20'):
                try:
                    month, day = int(pc[4:6]), int(pc[6:8])
                    if 1 <= month <= 12 and 1 <= day <= 31: continue 
                except: pass
            chart_no = pc
            break

        # --- 2. 성명 추출 (한글 소거법) ---
        korean_text = re.sub(r'[^가-힣]', '', line_compact)
        stopwords = [
            '마크로젠', '일반질환', '면역질환', '암특화', '온코체크', '토탈체크', '뷰티체크', '헬시체크',
            '주민번호', '차트번호', '검체번호', '결과코드', '처방코드', '수진자명', '환자성명', '검사항목',
            '검사명', '성명', '이름', '환자', '샘플명', '유전자', '바코드', '페이지', '리스트',
            '기관', '의뢰', '일반', '면역', '질환', '특화', '온코', '토탈', '뷰티', '체크', 
            '혈액', '소변', '종합', '차트', '주민', '처방', '코드', '확인', '필요', '목록', '담당',
            '부서', '접수', '검체', '채취', '동의', '의사', '병원', '센터', '의원', '건강', '검진',
            '수진', '나이', '성별', '연령', '종목', '항목', '기타', '비고', '일자', '시간', '출력', '발행', '남성', '여성'
        ]
        
        for stop in stopwords:
            korean_text = korean_text.replace(stop, '')
            
        name = "이름확인필요"
        if 2 <= len(korean_text) <= 4:
            name = korean_text
        elif len(korean_text) > 4:
            if korean_text[:2] in ['남궁', '황보', '제갈', '사공', '선우', '독고']:
                name = korean_text[:4]
            else:
                name = korean_text[:3]
                
        return name, chart_no

    def process_ocr_line(self, line):
        if len(line.strip()) < 10: return []
        
        name, chart_no = self.extract_refined_info(line)
        line_compact = line.replace(' ', '')
        
        barcodes = [n for n in re.findall(r'\d+', line_compact) if len(n) == 10 and n.startswith('26')]
        order_codes = list(dict.fromkeys(re.findall(r'(LZCMARC(?:AUTO|MB|FB|20[12]))', line_compact)))
        keys = re.findall(r'(2026\d{8})', line_compact)
        
        if not chart_no and not barcodes:
            return []
            
        jumin = ""
        jumin_match = re.search(r'(\d{6})-([1-8])', line)
        if jumin_match:
            jumin = f"{jumin_match.group(1)}-{jumin_match.group(2)}******"
            
        results = []
        num_tests = max(len(barcodes), len(order_codes), 1)
        
        for idx in range(num_tests):
            bc = barcodes[idx] if idx < len(barcodes) else ""
            oc = order_codes[idx] if idx < len(order_codes) else "LZCMARC202"
            kv = keys[idx] if idx < len(keys) else ""
            
            results.append({
                "성명": name, "차트번호": chart_no, "샘플명": "유전자", "검체": "EDT CEB",
                "검사명": ORDER_CODE_MAP.get(oc, "검사명 확인필요"),
                "바코드": bc, "처방코드": oc, "결과코드": oc,
                "주민번호": jumin, "Dept": "CPNW", "Room": "-", "Doct": "HC99", "Key": kv, "구분": "HC"
            })
        return results

# --- [로직 추가] OCR 결과물을 최종 업로드 서식으로 보정 ---
def process_wooriwon_logic(df):
    res_list = []
    last_valid_name = ""
    last_valid_chart = ""

    for idx, row in df.iterrows():
        # 주민번호 파싱
        ssn = str(row.get('주민번호', '')).strip()
        birth_year = ""
        gender = ""
        if len(ssn) >= 7 and ssn[6] == '-':
            birth_part = ssn[:6]
            gender_part = ssn[7]
            if gender_part in ['1', '2', '5', '6']: birth_year = "19" + birth_part[:2]
            elif gender_part in ['3', '4', '7', '8']: birth_year = "20" + birth_part[:2]
            gender = 'M' if gender_part in ['1', '3', '5', '7'] else 'F'
            
        # Key 값에서 채혈일 추출
        key_val = str(row.get('Key', '')).strip()
        date_val = f"{key_val[:4]}-{key_val[4:6]}-{key_val[6:8]}" if len(key_val) >= 8 else key_val
        
        name = str(row.get('성명', '')).strip()
        chart_no = str(row.get('차트번호', '')).replace('.0', '').strip()
        if chart_no == 'nan': chart_no = ''
        
        is_suspicious = False
        
        # 다중 검사 및 OCR 에러 보정 로직
        if name in ['이름확인필요', 'nan', ''] or name in ['구전자', '국전자', '종여', '종남', '주전자']:
            # 바로 윗줄과 주민번호가 같다면 동일인의 다중검사이므로 윗줄 정보 복사
            if idx > 0 and ssn == str(df.iloc[idx-1].get('주민번호', '')).strip():
                name = last_valid_name
                if not chart_no: chart_no = last_valid_chart
            else:
                is_suspicious = True
        
        if not is_suspicious:
            last_valid_name = name
            last_valid_chart = chart_no

        new_row = {
            '병원명': '우리원의료재단',
            '성명': name,
            '성명_수정필요': 'O' if is_suspicious else '',
            '차트번호': chart_no,
            '검체 바코드 번호(Kit ID)': str(row.get('바코드', '')).strip(),
            '샘플상태': 'Blood',
            '출생연도': birth_year,
            '성별': gender,
            '검체채취일': date_val,
            '동의서 및 의뢰서': date_val,
            '서비스코드': str(row.get('검사명', '')).strip(),
            '주민번호_원본': ssn,
        }
        res_list.append(new_row)
        
    return pd.DataFrame(res_list)

# --- [UI] 메인 화면 ---
st.set_page_config(page_title="우리원 OCR 추출기", layout="wide")
st.title("📄 우리원 OCR 자동화 시스템 (추출 및 서식변환 통합)")
st.info("💡 PDF를 업로드하면 내용 추출부터 최종 엑셀 서식 변환까지 한 번에 처리됩니다.")

uploaded_file = st.file_uploader("분석할 PDF 의뢰서를 업로드하세요", type=['pdf'])

if uploaded_file:
    if st.button("🚀 정밀 분석 시작"):
        converter = PDFConverter()
        doc = fitz.open(stream=uploaded_file.read(), filetype="pdf")
        all_rows = []
        
        progress_text = st.empty()
        progress_bar = st.progress(0)
        
        for i, page in enumerate(doc):
            progress_text.text(f"🔍 페이지 스캔 중... ({i+1}/{len(doc)})")
            pix = page.get_pixmap(matrix=fitz.Matrix(3.5, 3.5), alpha=False)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            img = ImageEnhance.Contrast(ImageOps.grayscale(img)).enhance(2.5)
            
            page_text = pytesseract.image_to_string(img, lang='kor+eng', config='--oem 3 --psm 6')
            for line in page_text.split('\n'):
                all_rows.extend(converter.process_ocr_line(line))
            progress_bar.progress((i + 1) / len(doc))

        if all_rows:
            # 1차: 원본 데이터프레임 생성 및 중복 제거
            raw_df = pd.DataFrame(all_rows)
            raw_df = raw_df.drop_duplicates(subset=['차트번호', '바코드', '검사명'], keep='first')
            
            # 2차: 최종 서식으로 데이터 보정 및 변환
            final_df = process_wooriwon_logic(raw_df)
            
            progress_text.empty()
            
            # 오류 의심 건수 체크
            suspicious_count = len(final_df[final_df['성명_수정필요'] == 'O'])
            if suspicious_count > 0:
                st.warning(f"⚠️ 글씨 뭉개짐으로 성명 수동 확인이 필요한 항목이 {suspicious_count}건 있습니다. 엑셀을 다운로드 후 '성명_수정필요' 열을 확인해주세요.")
            else:
                st.success(f"✅ 분석 및 변환 완료! 총 {final_df['차트번호'].nunique()}명의 환자 데이터를 추출했습니다.")
                
            st.dataframe(final_df)
            
            # 엑셀 파일 생성
            out = io.BytesIO()
            with pd.ExcelWriter(out, engine='openpyxl') as writer:
                final_df.to_excel(writer, index=False)
                
                # 중복 바코드 셀 색상 하이라이트 적용
                wb = writer.book
                ws = wb.active
                sky_blue = PatternFill(start_color="B3E5FC", end_color="B3E5FC", fill_type="solid")
                multi_bc = final_df[final_df.duplicated('검체 바코드 번호(Kit ID)', keep=False)]['검체 바코드 번호(Kit ID)'].unique()
                
                for idx, row in final_df.iterrows():
                    bc_val = row['검체 바코드 번호(Kit ID)']
                    if bc_val in multi_bc and bc_val != "":
                        for c in range(1, len(final_df.columns) + 1):
                            ws.cell(row=idx+2, column=c).fill = sky_blue
                            
            st.download_button(
                label="📥 최종 결과 엑셀 다운로드", 
                data=out.getvalue(), 
                file_name=f"우리원_최종변환_{datetime.now().strftime('%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )